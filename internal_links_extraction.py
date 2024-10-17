import requests
from bs4 import BeautifulSoup
import openpyxl

# Function to scrape internal links for a single article
def scrape_internal_links(url):
    try:
        # Fetch the webpage content
        response = requests.get(url)
        
        # Check if the request was successful
        if response.status_code != 200:
            print(f"Failed to retrieve the webpage for {url}. Status code: {response.status_code}")
            return []
        
        # Parse the HTML content using BeautifulSoup
        soup = BeautifulSoup(response.content, 'html.parser')
        
        # Find the <article> tag content
        article = soup.find('article')
        if not article:
            print(f"No <article> tag found for {url}.")
            return []
        
        # Exclude the div with id="mh-related-posts" if it exists
        related_posts_div = article.find('div', id='mh-related-posts')
        if related_posts_div:
            related_posts_div.decompose()  # Remove the div and all its children
        
        # Extract all anchor (<a>) tags within the article
        links = article.find_all('a')
        
        # Prepare data for returning: Extract href and anchor text
        link_data = []
        for link in links:
            href = link.get('href')
            anchor_text = link.get_text(strip=True)
            link_data.append([href, anchor_text])
        
        return link_data
    
    except Exception as e:
        print(f"An error occurred while processing {url}: {e}")
        return []

# Function to read URLs from the input .xlsx, scrape internal links, and write the output to a new .xlsx
def scrape_links_from_xlsx(input_xlsx, output_xlsx):
    # Load the input Excel workbook and sheet
    wb_input = openpyxl.load_workbook(input_xlsx)
    ws_input = wb_input.active  # Assuming the URLs are in the first sheet
    
    # Create a new workbook for output
    wb_output = openpyxl.Workbook()
    ws_output = wb_output.active
    ws_output.title = "Scraped Links"
    
    # Write header to output Excel
    ws_output.append(['URL', 'Internal Links (Comma Separated)', 'Anchor Texts (Comma Separated)'])
    
    # Loop through each row in the input Excel (assuming URLs are in the first column)
    for row in ws_input.iter_rows(min_row=2, max_col=1, values_only=True):  # Skipping header (min_row=2)
        url = row[0]  # Get the URL from the first column
        
        if url:
            # Scrape internal links for this URL
            link_data = scrape_internal_links(url)
            
            if link_data:
                # Extract URLs and anchor texts separately
                internal_links = [data[0] for data in link_data]
                anchor_texts = [data[1] for data in link_data]
                
                # Write the scraped links and anchor texts to the output Excel
                ws_output.append([
                    url,  # URL from input
                    ', '.join(internal_links),  # Comma-separated internal links
                    ', '.join(anchor_texts)  # Comma-separated anchor texts
                ])
            else:
                # If no data was scraped, leave blank
                ws_output.append([url, '', ''])
    
    # Save the output Excel workbook
    wb_output.save(output_xlsx)
    print(f"Scraping completed. Data saved to {output_xlsx}")

# Example usage
input_xlsx = "article_urls.xlsx"  # Input file containing URLs
output_xlsx = "scraped_article_links.xlsx"  # Output file for scraped data
scrape_links_from_xlsx(input_xlsx, output_xlsx)
